from collections import OrderedDict
from datetime import datetime
from typing import Any, Dict, List, Tuple
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd


class ExcelTableProcessor:
    def __init__(self, file_path: str) -> None:
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(self.file_path, data_only=True)

    def process_sheets(self) -> Dict[str, List[str]]:
        results = OrderedDict()
        for sheetname in self.wb.sheetnames:
            sheet = self.wb[sheetname]
            tables = self.identify_tables(sheet)
            results[sheetname] = tables
        return results

    def format_cell_value(self, cell: Cell) -> str:
        """Format the value of a cell based on its number_format."""
        value = cell.value
        if value is None:
            return ""

        if cell.is_date:
            # Date formatting
            date_format = {
                'd-mmm-yy': '%d-%b-%y',
                'mmm-yy': '%b-%y',
                '"FYE"\\ mmmyy': 'FYE %b%y',
                '"YTD"\\ mmmyy': 'YTD %b%y'
            }.get(cell.number_format, '%Y-%m-%d')
            return value.strftime(date_format).upper()
        elif isinstance(value, (int, float)):
            # Number formatting
            if cell.number_format in ['"$"#,##0_);\("$"#,##0\)']:
                formatted_value = f"${value:,.0f}" if value >= 0 else f"(${abs(value):,.0f})"
            elif cell.number_format == '"$"* #,##0.00\\ _€':
                #formatted_value = f"${value:,.2f} €"

                # Adjust for currency symbol and correct alignment
                formatted_value = f"${value:,.2f}"
                # Align the string to the right for a total of 15 characters width
                formatted_value = formatted_value.rjust(15)
            elif cell.number_format == '_("$"* #,##0_);_("$"* \(#,##0\);_("$"* "-"??_);_(@_)':
                # Handle complex currency format with alignment
                if value > 0:
                    formatted_value = f"${value:,.0f}"
                elif value < 0:
                    formatted_value = f"(${abs(value):,.0f})"
                else:  # Assuming zero or other cases
                    formatted_value = f"$ -"
                formatted_value = f"{formatted_value:>15}"  # Align right within 15 character width
            elif cell.number_format == '0%':
                formatted_value = f"{value:.0%}"
            elif cell.number_format == '0.0%':
                formatted_value = f"{value:.1%}"
            else:
                formatted_value = str(value)
            return formatted_value
        else:
            # Convert value to string and preserve leading spaces using HTML non-breaking spaces
            value_str = str(value)
            leading_spaces = len(value_str) - len(value_str.lstrip(' '))
            if leading_spaces:
                preserved_spaces = '&nbsp;' * leading_spaces
                value_str = preserved_spaces + value_str.lstrip(' ')
            return value_str

    def has_required_left_border(self, cell: Cell) -> bool:
        """Check if a cell has the required left border."""
        return cell.border.left.style in ['thin', 'medium', 'thick']

    def has_required_top_border(self, cell: Cell) -> bool:
        """Check if a cell has the required top border."""
        return cell.border.top.style in ['thin', 'medium', 'thick']

    def is_first_header_cell(self, cell: Cell) -> bool:
        """Check if a cell meets the first header characteristics."""
        return self.has_required_left_border(cell) and self.has_required_top_border(cell)

    def find_table(self, sheet: Worksheet, start_row: int, start_col: int) -> Tuple[int, int, int, int]:
        """Identify the width and height of the table starting from a header cell."""
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Identify table width
        n = 1
        while start_col + n <= max_col and self.has_required_top_border(sheet.cell(row=start_row, column=start_col + n)):
            n += 1

        # Identify table height
        m = 1
        while start_row + m <= max_row and self.has_required_left_border(sheet.cell(row=start_row + m, column=start_col)):
            m += 1

        return (start_row, start_col, start_row + m - 1, start_col + n - 1)

    def convert_table_to_markdown(self, sheet: Worksheet, table_range: Tuple[int, int, int, int]) -> str:
        """Convert the specified range of cells into a Markdown formatted table."""
        data = []
        for r in range(table_range[0], table_range[2] + 1):
            row_data = [self.format_cell_value(sheet.cell(row=r, column=c)) for c in range(table_range[1], table_range[3] + 1)]
            data.append(row_data)

        df = pd.DataFrame(data[1:], columns=data[0])
        return df.to_markdown(index=False)

    def identify_tables(self, sheet: Worksheet) -> List[str]:
        """Scan the worksheet for tables and return a list of Markdown formatted tables."""
        sheet = self.wb.active
        max_row = sheet.max_row
        max_col = sheet.max_column
        markdown_tables = []
        processed_cells = set()

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if (row, col) not in processed_cells and self.is_first_header_cell(cell):
                    table_range = self.find_table(sheet, row, col)
                    markdown_tables.append(self.convert_table_to_markdown(sheet, table_range))
                    # Mark cells as processed
                    for r in range(table_range[0], table_range[2] + 1):
                        for c in range(table_range[1], table_range[3] + 1):
                            processed_cells.add((r, c))
        return markdown_tables
